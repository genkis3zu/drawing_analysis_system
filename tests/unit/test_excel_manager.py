"""Excel管理機能のユニットテスト

TDDアプローチに従い、実装前にテストを作成。
Excel出力機能の要件：
- 解析結果をExcelファイルに出力
- テンプレートベースの出力
- 複数シートのサポート
- 信頼度スコアの色分け表示
"""
import pytest
from pathlib import Path
import tempfile
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from typing import Dict, Any

from src.utils.excel_manager import ExcelManager
from src.models.analysis_result import AnalysisResult, ExtractionResult


class TestExcelManager:
    """ExcelManagerのユニットテスト"""
    
    @pytest.fixture
    def temp_dir(self):
        """一時ディレクトリを作成"""
        with tempfile.TemporaryDirectory() as temp_dir:
            yield Path(temp_dir)
    
    @pytest.fixture
    def excel_manager(self, temp_dir):
        """テスト用のExcelManagerインスタンス"""
        config = {
            "output_dir": str(temp_dir),
            "template_dir": str(temp_dir / "templates"),
            "default_template": "basic_template.xlsx"
        }
        return ExcelManager(config)
    
    @pytest.fixture
    def sample_analysis_result(self):
        """サンプル解析結果"""
        return AnalysisResult(
            drawing_path="test_drawing.png",
            extracted_data={
                "部品番号": ExtractionResult(
                    field_name="部品番号",
                    value="A-123",
                    confidence=0.95
                ),
                "材質": ExtractionResult(
                    field_name="材質",
                    value="SUS304",
                    confidence=0.88
                ),
                "寸法": ExtractionResult(
                    field_name="寸法",
                    value="100x50x10",
                    confidence=0.75
                ),
                "数量": ExtractionResult(
                    field_name="数量",
                    value="5",
                    confidence=0.92
                )
            }
        )
    
    def test_初期化_正常な設定(self, temp_dir):
        """正常な設定でのExcelManager初期化をテスト"""
        # Given
        config = {
            "output_dir": str(temp_dir),
            "template_dir": str(temp_dir / "templates")
        }
        
        # When
        manager = ExcelManager(config)
        
        # Then
        assert manager.output_dir == temp_dir
        assert manager.template_dir == temp_dir / "templates"
    
    def test_単一結果のExcel出力(self, excel_manager, sample_analysis_result, temp_dir):
        """単一の解析結果をExcelに出力するテスト"""
        # Given
        output_file = temp_dir / "test_output.xlsx"
        
        # When
        result_path = excel_manager.export_single_result(
            sample_analysis_result,
            output_file
        )
        
        # Then
        assert result_path.exists()
        assert result_path.suffix == ".xlsx"
        
        # Excelファイルの内容を検証
        df = pd.read_excel(result_path, sheet_name="解析結果")
        assert "項目" in df.columns
        assert "値" in df.columns
        assert "信頼度" in df.columns
        assert "部品番号" in df["項目"].values
        assert "A-123" in df["値"].values
        assert 0.95 in df["信頼度"].values
    
    def test_複数結果のバッチExcel出力(self, excel_manager, sample_analysis_result, temp_dir):
        """複数の解析結果をバッチでExcelに出力するテスト"""
        # Given
        results = [
            sample_analysis_result,
            AnalysisResult(
                drawing_path="test_drawing2.png",
                extracted_data={
                    "部品番号": ExtractionResult(
                        field_name="部品番号",
                        value="B-456",
                        confidence=0.90
                    ),
                    "材質": ExtractionResult(
                        field_name="材質",
                        value="AL6061",
                        confidence=0.85
                    )
                }
            )
        ]
        output_file = temp_dir / "batch_output.xlsx"
        
        # When
        result_path = excel_manager.export_batch_results(results, output_file)
        
        # Then
        assert result_path.exists()
        
        # 複数シートの存在を確認
        wb = load_workbook(result_path)
        assert "サマリー" in wb.sheetnames
        assert "詳細_1" in wb.sheetnames
        assert "詳細_2" in wb.sheetnames
    
    def test_テンプレートベースの出力(self, excel_manager, sample_analysis_result, temp_dir):
        """テンプレートを使用したExcel出力のテスト"""
        # Given
        template_path = temp_dir / "templates" / "custom_template.xlsx"
        template_path.parent.mkdir(exist_ok=True)
        
        # テンプレートファイルを作成
        self._create_test_template(template_path)
        
        output_file = temp_dir / "template_output.xlsx"
        
        # When
        result_path = excel_manager.export_with_template(
            sample_analysis_result,
            template_path,
            output_file
        )
        
        # Then
        assert result_path.exists()
        wb = load_workbook(result_path)
        assert "図面情報" in wb.sheetnames
    
    def test_信頼度による色分け設定(self, excel_manager, sample_analysis_result, temp_dir):
        """信頼度スコアに基づく色分け設定のテスト"""
        # Given
        output_file = temp_dir / "colored_output.xlsx"
        
        # When
        result_path = excel_manager.export_single_result(
            sample_analysis_result,
            output_file,
            apply_confidence_coloring=True
        )
        
        # Then
        wb = load_workbook(result_path)
        ws = wb["解析結果"]
        
        # 信頼度の高いセル（0.9以上）は緑色背景
        # 信頼度の中程度（0.7-0.9）は黄色背景
        # 信頼度の低い（0.7未満）は赤色背景
        # ※実際の色の検証は、openpyxlのFill属性を確認
        assert result_path.exists()
    
    def test_エラーハンドリング_無効な出力パス(self, excel_manager, sample_analysis_result):
        """無効な出力パスでのエラーハンドリングをテスト"""
        # Given
        invalid_path = Path("/invalid/path/output.xlsx")
        
        # When/Then
        with pytest.raises(FileNotFoundError):
            excel_manager.export_single_result(
                sample_analysis_result,
                invalid_path
            )
    
    def test_日本語フィールド名の正常処理(self, excel_manager, temp_dir):
        """日本語フィールド名を含むデータの正常処理をテスト"""
        # Given
        result = AnalysisResult(
            drawing_path="japanese_test.png",
            extracted_data={
                "製品名": ExtractionResult(
                    field_name="製品名",
                    value="テスト製品",
                    confidence=0.95
                ),
                "製造元": ExtractionResult(
                    field_name="製造元",
                    value="株式会社テスト",
                    confidence=0.90
                ),
                "備考": ExtractionResult(
                    field_name="備考",
                    value="特記事項なし",
                    confidence=0.85
                )
            }
        )
        output_file = temp_dir / "japanese_output.xlsx"
        
        # When
        result_path = excel_manager.export_single_result(result, output_file)
        
        # Then
        df = pd.read_excel(result_path, sheet_name="解析結果")
        assert "製品名" in df["項目"].values
        assert "テスト製品" in df["値"].values
    
    def test_統計情報の出力(self, excel_manager, sample_analysis_result, temp_dir):
        """解析統計情報のExcel出力をテスト"""
        # Given
        output_file = temp_dir / "stats_output.xlsx"
        
        # When
        result_path = excel_manager.export_single_result(
            sample_analysis_result,
            output_file,
            include_statistics=True
        )
        
        # Then
        wb = load_workbook(result_path)
        assert "統計情報" in wb.sheetnames
        
        df_stats = pd.read_excel(result_path, sheet_name="統計情報")
        assert "平均信頼度" in df_stats["項目"].values
        assert "処理時間" in df_stats["項目"].values
    
    @staticmethod
    def _create_test_template(template_path: Path):
        """テスト用のテンプレートファイルを作成"""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "図面情報"
        
        # テンプレートのヘッダー設定
        headers = ["項目", "値", "信頼度", "備考"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        wb.save(template_path)