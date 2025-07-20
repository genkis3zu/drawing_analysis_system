# src/utils/excel_manager.py

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import logging
from datetime import datetime
import io

from ..models.analysis_result import AnalysisResult

class ExcelManager:
    """エクセルファイル管理クラス"""
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """初期化
        
        Args:
            config: 設定辞書。以下のキーを含む：
                - output_dir: 出力ディレクトリ（必須）
                - template_dir: テンプレートディレクトリ（必須）
                - default_template: デフォルトテンプレート名（オプション）
        """
        if config is None:
            config = {
                "output_dir": "data/excel_output",
                "template_dir": "data/excel_templates"
            }
        
        self.output_dir = Path(config["output_dir"])
        self.template_dir = Path(config["template_dir"])
        self.default_template = config.get("default_template", "basic_template.xlsx")
        
        # ディレクトリ作成
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.template_dir.mkdir(parents=True, exist_ok=True)
        
        self.logger = logging.getLogger(__name__)
    
    def export_single_result(
        self,
        analysis_result: AnalysisResult,
        output_path: Path,
        apply_confidence_coloring: bool = False,
        include_statistics: bool = False
    ) -> Path:
        """単一の解析結果をExcelに出力する
        
        Args:
            analysis_result: 解析結果
            output_path: 出力ファイルパス
            apply_confidence_coloring: 信頼度による色分けを適用するか
            include_statistics: 統計情報を含めるか
            
        Returns:
            Path: 作成されたExcelファイルのパス
            
        Raises:
            FileNotFoundError: 出力パスの親ディレクトリが存在しない場合
        """
        # 出力パスの親ディレクトリが存在するか確認
        if not output_path.parent.exists():
            raise FileNotFoundError(f"出力ディレクトリが存在しません: {output_path.parent}")
        
        # ワークブック作成
        wb = openpyxl.Workbook()
        
        # 解析結果シート作成
        self._create_results_sheet_v2(wb, analysis_result, apply_confidence_coloring)
        
        # 統計情報シート作成（オプション）
        if include_statistics:
            self._create_statistics_sheet(wb, analysis_result)
        
        # 保存
        wb.save(output_path)
        
        self.logger.info(f"Excel出力完了: {output_path}")
        return output_path
    
    def create_analysis_report(self, analysis_result: AnalysisResult, output_path: str = None) -> str:
        """解析結果からエクセルレポートを作成"""
        
        try:
            # 出力パス決定
            if output_path is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"analysis_report_{timestamp}.xlsx"
                output_path = self.template_dir.parent / "excel_output" / filename
                output_path.parent.mkdir(parents=True, exist_ok=True)
            
            # ワークブック作成
            wb = openpyxl.Workbook()
            
            # シート1: 解析結果
            self._create_results_sheet(wb, analysis_result)
            
            # シート2: メタデータ
            self._create_metadata_sheet(wb, analysis_result)
            
            # シート3: 品質評価
            self._create_quality_sheet(wb, analysis_result)
            
            # 保存
            wb.save(output_path)
            
            self.logger.info(f"エクセルレポート作成完了: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"エクセルレポート作成エラー: {e}")
            raise
    
    def _create_results_sheet(self, workbook: openpyxl.Workbook, analysis_result: AnalysisResult):
        """解析結果シートを作成"""
        
        ws = workbook.active
        ws.title = "解析結果"
        
        # ヘッダー設定
        headers = ["フィールド名", "抽出値", "信頼度", "位置", "抽出方法", "バリデーション", "備考"]
        
        # ヘッダー行作成
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # データ行作成
        row = 2
        for field_name, extraction_result in analysis_result.extracted_data.items():
            
            # フィールド名
            ws.cell(row=row, column=1, value=field_name)
            
            # 抽出値
            value_cell = ws.cell(row=row, column=2, value=str(extraction_result.value))
            
            # 信頼度
            confidence_cell = ws.cell(row=row, column=3, value=extraction_result.confidence)
            confidence_cell.number_format = "0%"
            
            # 信頼度による色分け
            if extraction_result.confidence >= 0.8:
                value_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif extraction_result.confidence >= 0.6:
                value_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                value_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # 位置情報
            position_str = ""
            if extraction_result.position:
                pos = extraction_result.position
                position_str = f"({pos.get('x', 0)}, {pos.get('y', 0)})"
            ws.cell(row=row, column=4, value=position_str)
            
            # 抽出方法
            ws.cell(row=row, column=5, value=extraction_result.extraction_method.value)
            
            # バリデーション状態
            validation_cell = ws.cell(row=row, column=6, value=extraction_result.validation_status)
            if extraction_result.validation_status == "valid":
                validation_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif extraction_result.validation_status == "invalid":
                validation_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # 備考
            ws.cell(row=row, column=7, value=extraction_result.notes or "")
            
            row += 1
        
        # 列幅調整
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # 罫線追加
        self._add_borders(ws, 1, 1, row - 1, len(headers))
    
    def _create_metadata_sheet(self, workbook: openpyxl.Workbook, analysis_result: AnalysisResult):
        """メタデータシートを作成"""
        
        ws = workbook.create_sheet(title="メタデータ")
        
        metadata_items = [
            ("解析ID", analysis_result.result_id),
            ("図面パス", analysis_result.drawing_path),
            ("解析日時", analysis_result.created_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("テンプレートID", analysis_result.template_id or "なし"),
            ("製品タイプ", analysis_result.product_type or "不明"),
            ("総合信頼度", f"{analysis_result.confidence_score:.1%}"),
            ("抽出フィールド数", len(analysis_result.extracted_data)),
            ("高信頼度フィールド数", analysis_result.quality_metrics.high_confidence_fields),
            ("処理時間", f"{analysis_result.processing_metrics.processing_time:.2f}秒"),
            ("APIコール数", analysis_result.processing_metrics.api_calls),
        ]
        
        # A4情報追加
        if analysis_result.a4_info:
            a4_info = analysis_result.a4_info
            metadata_items.extend([
                ("", ""),  # 空行
                ("=== A4図面情報 ===", ""),
                ("図面向き", a4_info.get('orientation', '不明')),
                ("A4サイズ", "✓" if a4_info.get('is_valid_a4') else "✗"),
                ("DPI", a4_info.get('dpi', 0)),
                ("画像サイズ", f"{a4_info.get('width', 0)}×{a4_info.get('height', 0)}"),
            ])
        
        # メタデータ項目を設定
        for row, (key, value) in enumerate(metadata_items, 1):
            key_cell = ws.cell(row=row, column=1, value=key)
            value_cell = ws.cell(row=row, column=2, value=value)
            
            # ヘッダー行のスタイル
            if key.startswith("==="):
                key_cell.font = Font(bold=True)
                key_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            elif key:
                key_cell.font = Font(bold=True)
        
        # 列幅調整
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
    
    def _create_quality_sheet(self, workbook: openpyxl.Workbook, analysis_result: AnalysisResult):
        """品質評価シートを作成"""
        
        ws = workbook.create_sheet(title="品質評価")
        
        # 品質メトリクス
        quality_items = [
            ("総合品質評価", ""),
            ("総合信頼度", f"{analysis_result.quality_metrics.overall_confidence:.1%}"),
            ("高信頼度フィールド率", 
             f"{analysis_result.quality_metrics.high_confidence_fields}/{analysis_result.quality_metrics.total_fields}"),
            ("バリデーション通過率", f"{analysis_result.quality_metrics.validation_pass_rate:.1%}"),
            ("抽出完全性", f"{analysis_result.quality_metrics.extraction_completeness:.1%}"),
            ("", ""),
            ("処理パフォーマンス", ""),
            ("総処理時間", f"{analysis_result.processing_metrics.processing_time:.2f}秒"),
            ("画像前処理時間", f"{analysis_result.processing_metrics.image_preprocessing_time:.2f}秒"),
            ("AI解析時間", f"{analysis_result.processing_metrics.ai_analysis_time:.2f}秒"),
            ("後処理時間", f"{analysis_result.processing_metrics.post_processing_time:.2f}秒"),
        ]
        
        for row, (key, value) in enumerate(quality_items, 1):
            key_cell = ws.cell(row=row, column=1, value=key)
            value_cell = ws.cell(row=row, column=2, value=value)
            
            if key and not value:  # ヘッダー行
                key_cell.font = Font(bold=True, size=12)
                key_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            elif key:
                key_cell.font = Font(bold=True)
        
        # フィールド別信頼度グラフ用データ
        ws.cell(row=len(quality_items) + 3, column=1, value="フィールド別信頼度").font = Font(bold=True, size=12)
        
        chart_row = len(quality_items) + 5
        ws.cell(row=chart_row, column=1, value="フィールド名").font = Font(bold=True)
        ws.cell(row=chart_row, column=2, value="信頼度").font = Font(bold=True)
        
        for i, (field_name, extraction_result) in enumerate(analysis_result.extracted_data.items(), 1):
            ws.cell(row=chart_row + i, column=1, value=field_name)
            confidence_cell = ws.cell(row=chart_row + i, column=2, value=extraction_result.confidence)
            confidence_cell.number_format = "0%"
        
        # 列幅調整
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def export_batch_results(
        self,
        analysis_results: List[AnalysisResult],
        output_path: Path
    ) -> Path:
        """複数の解析結果をバッチでExcelに出力する
        
        Args:
            analysis_results: 解析結果のリスト
            output_path: 出力ファイルパス
            
        Returns:
            Path: 作成されたExcelファイルのパス
        """
        wb = openpyxl.Workbook()
        
        # サマリーシート
        self._create_batch_summary_sheet(wb, analysis_results)
        
        # 詳細シート（結果ごと）
        for i, result in enumerate(analysis_results, 1):
            sheet_name = f"詳細_{i}"
            ws = wb.create_sheet(title=sheet_name)
            self._add_result_to_sheet(ws, result)
        
        # 保存
        wb.save(output_path)
        
        self.logger.info(f"バッチExcel出力完了: {output_path}")
        return output_path
    
    def export_with_template(
        self,
        analysis_result: AnalysisResult,
        template_path: Path,
        output_path: Path
    ) -> Path:
        """テンプレートを使用してExcelに出力する
        
        Args:
            analysis_result: 解析結果
            template_path: テンプレートファイルパス
            output_path: 出力ファイルパス
            
        Returns:
            Path: 作成されたExcelファイルのパス
        """
        # テンプレート読み込み
        wb = openpyxl.load_workbook(template_path)
        
        # 図面情報シートに出力
        if "図面情報" in wb.sheetnames:
            ws = wb["図面情報"]
            self._populate_template_sheet(ws, analysis_result)
        
        # 保存
        wb.save(output_path)
        
        self.logger.info(f"テンプレートベースExcel出力完了: {output_path}")
        return output_path
    
    def create_batch_report(self, analysis_results: List[AnalysisResult], output_path: str = None) -> str:
        """複数解析結果のバッチレポートを作成"""
        
        try:
            if output_path is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"batch_report_{timestamp}.xlsx"
                output_path = self.template_dir.parent / "excel_output" / filename
                output_path.parent.mkdir(parents=True, exist_ok=True)
            
            wb = openpyxl.Workbook()
            
            # サマリーシート
            self._create_batch_summary_sheet(wb, analysis_results)
            
            # 詳細シート（最大10件まで）
            for i, result in enumerate(analysis_results[:10], 1):
                sheet_name = f"詳細_{i}"
                try:
                    ws = wb.create_sheet(title=sheet_name)
                except Exception:
                    ws = wb.create_sheet(title=f"Sheet_{i}")
                
                # 詳細データを追加
                self._add_result_to_sheet(ws, result)
            
            # 保存
            wb.save(output_path)
            
            self.logger.info(f"バッチレポート作成完了: {output_path}")
            return str(output_path)
            
        except Exception as e:
            self.logger.error(f"バッチレポート作成エラー: {e}")
            raise
    
    def _add_borders(self, ws, start_row: int, start_col: int, end_row: int, end_col: int):
        """指定範囲に罫線を追加"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = thin_border
    
    def _create_batch_summary_sheet(self, workbook: openpyxl.Workbook, analysis_results: List[AnalysisResult]):
        """バッチサマリーシートを作成"""
        ws = workbook.active
        ws.title = "サマリー"
        
        # ヘッダー
        headers = ["No.", "図面名", "解析日時", "総合信頼度", "抽出フィールド数", "高信頼度フィールド数", "ステータス"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # データ行
        for row, result in enumerate(analysis_results, 2):
            ws.cell(row=row, column=1, value=row - 1)
            ws.cell(row=row, column=2, value=Path(result.drawing_path).name)
            ws.cell(row=row, column=3, value=result.created_at.strftime("%Y-%m-%d %H:%M"))
            
            confidence_cell = ws.cell(row=row, column=4, value=result.confidence_score)
            confidence_cell.number_format = "0%"
            
            ws.cell(row=row, column=5, value=len(result.extracted_data))
            ws.cell(row=row, column=6, value=result.quality_metrics.high_confidence_fields)
            
            status = "成功" if result.confidence_score >= 0.7 else "要確認"
            status_cell = ws.cell(row=row, column=7, value=status)
            if status == "成功":
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # 列幅調整
        column_widths = [5, 30, 20, 15, 20, 20, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # 罫線追加
        self._add_borders(ws, 1, 1, len(analysis_results) + 1, len(headers))
    
    def _add_result_to_sheet(self, ws, result: AnalysisResult):
        """シートに解析結果を追加"""
        # 基本情報
        ws.cell(row=1, column=1, value="図面名").font = Font(bold=True)
        ws.cell(row=1, column=2, value=Path(result.drawing_path).name)
        
        ws.cell(row=2, column=1, value="解析日時").font = Font(bold=True)
        ws.cell(row=2, column=2, value=result.created_at.strftime("%Y-%m-%d %H:%M:%S"))
        
        ws.cell(row=3, column=1, value="総合信頼度").font = Font(bold=True)
        confidence_cell = ws.cell(row=3, column=2, value=result.confidence_score)
        confidence_cell.number_format = "0%"
        
        # 抽出データ
        ws.cell(row=5, column=1, value="抽出データ").font = Font(bold=True, size=12)
        
        headers = ["フィールド名", "抽出値", "信頼度"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        row = 7
        for field_name, extraction_result in result.extracted_data.items():
            ws.cell(row=row, column=1, value=field_name)
            ws.cell(row=row, column=2, value=str(extraction_result.value))
            confidence_cell = ws.cell(row=row, column=3, value=extraction_result.confidence)
            confidence_cell.number_format = "0%"
            row += 1
        
        # 列幅調整
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
    
    def _create_results_sheet_v2(self, workbook: openpyxl.Workbook, analysis_result: AnalysisResult, apply_coloring: bool = False):
        """解析結果シートを作成（テストバージョン）"""
        ws = workbook.active
        ws.title = "解析結果"
        
        # ヘッダー設定
        headers = ["項目", "値", "信頼度"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # データ行作成
        row = 2
        for field_name, extraction_result in analysis_result.extracted_data.items():
            ws.cell(row=row, column=1, value=field_name)
            ws.cell(row=row, column=2, value=str(extraction_result.value))
            
            confidence_cell = ws.cell(row=row, column=3, value=extraction_result.confidence)
            confidence_cell.number_format = "0.00"
            
            # 信頼度による色分け（オプション）
            if apply_coloring:
                if extraction_result.confidence >= 0.9:
                    confidence_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif extraction_result.confidence >= 0.7:
                    confidence_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    confidence_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            row += 1
        
        # 列幅調整
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
    
    def _create_statistics_sheet(self, workbook: openpyxl.Workbook, analysis_result: AnalysisResult):
        """統計情報シートを作成"""
        ws = workbook.create_sheet(title="統計情報")
        
        # 統計データ作成
        statistics_items = [
            ("平均信頼度", f"{analysis_result.quality_metrics.overall_confidence:.2%}"),
            ("処理時間", f"{analysis_result.processing_metrics.processing_time:.2f}秒"),
            ("抽出フィールド数", str(len(analysis_result.extracted_data))),
            ("高信頼度フィールド数", str(analysis_result.quality_metrics.high_confidence_fields)),
            ("バリデーション通過率", f"{analysis_result.quality_metrics.validation_pass_rate:.2%}")
        ]
        
        # ヘッダー
        ws.cell(row=1, column=1, value="項目").font = Font(bold=True)
        ws.cell(row=1, column=2, value="値").font = Font(bold=True)
        
        # データ行
        for row, (key, value) in enumerate(statistics_items, 2):
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
        
        # 列幅調整
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _populate_template_sheet(self, ws, analysis_result: AnalysisResult):
        """テンプレートシートにデータを入力"""
        # テンプレートフォーマットに従ってデータを入力
        # ヘッダー行の次からデータを追加
        row = 2
        for field_name, extraction_result in analysis_result.extracted_data.items():
            # 既存の行にデータを挿入または新規追加
            ws.cell(row=row, column=1, value=field_name)
            ws.cell(row=row, column=2, value=str(extraction_result.value))
            ws.cell(row=row, column=3, value=extraction_result.confidence)
            ws.cell(row=row, column=4, value="")
            row += 1
